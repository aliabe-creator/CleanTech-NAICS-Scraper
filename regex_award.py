'''
Created on Sep 27, 2021

@author: Private
'''
from openpyxl import load_workbook
import re

wb = load_workbook("Oregon_AwardsSubawardsContractsSubcontracts/Truncated_ContractsSub.xlsx")
ws = wb.active

def general(s): #returns 1 if any regex in this section matches
    score = 0
    
    regex_terms = ['sustainab', 'green good', 'green technolog', 'green innov', 'eco[^ ]*innov', 'green manufac', 'green prod', 'pollut',
               'ecolabel', 'environ.* product declarat', 'EPD.*environ|environ.*EPD', 'environ.* prefer.* product', 'environ.* label']
    
    for term in regex_terms:
        yesorno = re.search(term, s)
        if yesorno:
            score = 1
    
    if score == 1:
        print('general hit')
            
    return score

def environ(s): #returns +1 if any regex in each section matches
    score = 0
    
    #all-purpose
    if (re.search('natur.* environ', s) or re.search('environ.* friend', s) or re.search('environment.* conserv', s) or re.search('biocompat', s) 
        or re.search('biodivers', s) or re.search('filter', s) or re.search('filtra', s) or re.search('synth.* gas', s) or re.search('regenerat', s)
        or re.search('recircul', s) or re.search('gasification', s) or re.search('gasifier', s) or re.search('fluidized clean gas', s)
        or re.search('gas cleaning', s)):
        
        score += 1
        print('environ - all purpose hit')
    
    #biotreatment
    if ((re.search('biogas.*', s) or re.search('bioreact.*', s) or re.search('polyolef.*', s) or re.search('biopolymer.*', s) 
        or re.search('disinfect.*', s) or re.search('biofilm.*', s) or re.search('biosens.*', s) or re.search('biosolid.*', s) 
        or re.search('caprolact.*', s) or ((re.search('ultraviol.*', s) or re.search('UV', s)) and (re.search('radiat.*', s) 
        or re.search('sol.*', s)))) and (re.search('bioremed.*', s) or re.search('biorecov.*', s) or (re.search('biolog.* treat.*', s)) 
        or re.search('biodegrad.*', s))):
        
        score += 1
        print('environ - biotreatment hit')
    
    #air pollution
    if ((re.search('air.* contr.*', s) or re.search('dust.* contr.*', s) or re.search('particular.* contr.*', s) 
        or re.search('air.* qual.*', s)) and re.search('pollut.*', s)):
        
        score += 1
        print('environ - air pollut hit')
    
    #environmental monitoring
    if  (re.search('environ.* monitor.*', s) and ((re.search('environ.* and instrument.*', s) 
        or re.search('environ.* and analys.*', s)) or re.search('life.*cycle analysis', s) or re.search('life cycle analys.*', s))):
        
        score += 1
        print('environ - monitoring hit')
    
    #marine pollution
    if re.search('marin.* control.*pollut|pollut.*marin.* control', s):
        score += 1
        print('environ - marine pollut hit')
    
    #noise and vib control
    if (re.search('nois.* abat.*', s) or re.search('nois.* reduc.*', s) or re.search('nois.* lessen.*', s)):
        score += 1
        print('environ - noise vib control hit')
    
    #land reclamation
    if (re.search('land', s) and (re.search('reclam.*', s) or re.search('remediat.*', s) or re.search('contamin.*', s))):
        score += 1
        print('environ - land reclamation hit')
        
    #waste management
    if (re.search('wast.*', s) or re.search('sewag.*', s) or re.search('inciner.*', s)):
        score += 1
        print('environ - waste manage hit')
        
    #water supply, treatment
    if ((re.search('slurr.*', s) or re.search('sludg.*', s) or re.search('aque.* solution.*', s) or re.search('wastewat.*', s) 
        or re.search('effluent.*', s) or re.search('sediment.*', s) or re.search('floccul.*', s) or re.search('detergen.*', s) 
        or re.search('coagul.*', s) or re.search('dioxin.*', s) or re.search('flow.* control.* dev.*', s) 
        or re.search('fluid commun.*', s) or re.search('high purit.*', s) or re.search('impur.*', s) or re.search('zeolit.*', s)) 
        and (re.search('water treat.*', s) or re.search('water purif.*', s) or re.search('water pollut.*', s))):
        
        score += 1
        print('environ - water supply hit')
    
    #recovery, recyling
    if (re.search('recycl.*', s) or re.search('compost.*', s) or re.search('stock process.*', s) or re.search('coal combust.*', s) 
        or re.search('remanufactur.*', s) or (re.search('coal', s) and re.search('PCC', s)) 
        or re.search('circulat.* fluid.* bed combust.*', s) or (re.search('combust.*', s) and re.search('CFBC', s))):
        
        score += 1
        print('environ - recycling hit')
        
    return score

def renew(s):
    score = 0
    
    #all-purpose
    if (re.search('renewabl.*', s) and (re.search('energ.*', s) or re.search('electric.*', s))):
        score += 1
        print('renew - all purpose hit')
    
    #wave, tidal
    if ((re.search('two basin schem.*', s) or re.search('wave.* energ.*', s) or re.search('tid.* energ.*', s)) and re.search('electric.*', s)):
        score += 1
        print('renew - wave, tidal hit')
    
    #biomass
    if (re.search('biomass.*', s) or re.search('enzymat.* hydrolys.*', s) or re.search('bio.*bas.* product.*', s)):
        score += 1
        print('renew - biomass hit')
    
    #wind
    if (re.search('wind power.*', s) or re.search('wind energ.*', s) or re.search('wind farm.*', s) or re.search('turbin.* and wind.*', s)):
        score += 1
        print('renew - wind hit')
    
    #geotherm
    if ((re.search('whole system.*', s) and re.search('geotherm.*', s)) or re.search('geotherm.*', s) or re.search('geoexchang.*', s)):
        score += 1
        print('renew - geotherm hit')
    
    #solar
    if (re.search('solar.*', s) and (re.search('ener.*', s) or re.search('linear fresnel sys.*', s) or re.search('electric.*', s) 
        or re.search('cell.*', s) or re.search('heat.*', s) or re.search('cool.*', s) or re.search('photovolt.*', s) 
        or re.search('PV', s) or re.search('cdte', s) or re.search('cadmium tellurid.*', s) or re.search('PVC-U', s) 
        or re.search('photoelectr.*', s) or re.search('photoactiv.*', s) or re.search('sol.*gel.* process.*', s) 
        or re.search('evacuat.* tub.*', s) or re.search('flat plate collect.*', s) or re.search('roof integr.* system.*', s))):
        
        score += 1
        print('renew - solar hit')
    
    return score

def elc(s):
    score = 0
    
    #all purp
    if (re.search('low carbon', s) or re.search('zero carbon', s) or re.search('no carbon', s) or re.search('0 carbon', s) 
        or re.search('low.*carbon', s) or re.search('zero.*carbon', s) or re.search('no.*carbon', s)):
        
        score += 1
        print('elc - all purpose hit')
    
    #alt fuel vehicle
    if (re.search('electric.* vehic.*', s) or re.search('hybrid vehic.*', s) or re.search('electric.* motor.*', s) 
        or re.search('hybrid motor.*', s) or re.search('hybrid driv.*', s) or re.search('electric.* car.*', s)
        or re.search('hybrid car.*', s) or re.search('electric.* machin.*', s) or re.search('electric.* auto.*', s) 
        or re.search('hybrid auto.*', s) or re.search('yaw.* rat.* sens.*', s)):
        
        score += 1
        print('elc - alt fuel veh hit')
    
    #alt fuels
    if (re.search('alternat.* fuel.*', s) or re.search('mainstream.* fuel.*', s) or re.search('fuel cell.*', s) or re.search('nuclear powe.*', s) 
        or re.search('nuclear stat.*', s) or re.search('nuclear plant.*', s) or re.search('nuclear energ.*', s) 
        or re.search('nuclear and electric.*', s) or re.search('nuclear fuel.*', s) or re.search('fuel.* process.*', s) 
        or re.search('porous.* struct.*', s) or re.search('porous.* substrat.*', s) or re.search('solid.* oxid.* fuel.*', s) 
        or re.search('Fischer.*Tropsch.*', s) or re.search('refus.* deriv.* fuel.*', s) or re.search('refus.*deriv.* fuel.*', s) 
        or (re.search('fuel.*', s) and re.search('biotech.*', s) and (re.search('ethanol.*', s) or re.search('hydrogen.*', s))) 
        or re.search('bio.*fuel.*', s) or re.search('synthetic fuel', s) or re.search('combined heat and power', s) 
        or re.search('synth.* gas.*', s) or re.search('syngas', s)):
        
        score += 1
        print('elc - alt fuel hit')
    
    #electrochem
    if (re.search('electrochem.* cell.*', s) or re.search('electrochem.* fuel.*', s) or re.search('membran.* electrod.*', s) 
        or re.search('ion.* exchang.* membran.*', s) or re.search('ion.*exchang.* membran.*', s) or re.search('electrolyt.* cell.*', s) 
        or re.search('catalyt.* convers.*', s) or re.search('solid.* separat.*', s) or re.search('membran.* separat.*', s) 
        or re.search('ion.* exchang.* resin.*', s) or re.search('ion.*exchang.* resin.*', s) or re.search('proton.* exchang.* membra.*', s) 
        or re.search('proton.*exchang.* membra.*', s) or re.search('cataly.* reduc.*', s) or re.search('electrod.* membran.*', s) 
        or re.search('therm.* engin.*', s)):
        
        score += 1
        print('elc - electrochem hit')
    
    #battery
    if ((re.search('batter.*', s) or re.search('accumul.*', s)) and (re.search('charg.*', s) or re.search('rechar.*', s) 
        or re.search('turbocharg.*', s) or re.search('high capacit.*', s) or re.search('rapid charg.*', s) or re.search('long life', s) 
        or re.search('ultra.*', s) or re.search('solar', s) or re.search('no lead', s) or re.search('no mercury', s) 
        or re.search('no cadmium', s) or re.search('lithium.*ion.*', s) or re.search('lithium.* ion.*', s) or re.search('Li.*ion', s))):
        
        score += 1
        print('elc - battery hit')
    
    #addl energy sources
    if (re.search('addition.* energ.* sourc.*', s) or re.search('addition.* sourc.* of ener.*', s)):
        score += 1
        print('elc - addl energy hit')
    
    #carbon capture
    if ((re.search('carbon', s) and re.search('captu.*', s)) or (re.search('carbon', s) and re.search('stor.*', s)) 
        or re.search('carbon dioxid.*', s) or re.search('CO2', s)):
        
        score += 1
        print('elc - carbon capture hit')
    
    #energy manage
    if (re.search('ener.* sav.*', s) or re.search('ener.* effic.*', s) or re.search('energ.*effic.*', s) or re.search('energ.*sav.*', s) 
        or re.search('light.* emit.* diod.*', s) or re.search("^\W*LED", s) or re.search('organic LED', s) or re.search('OLED', s) 
        or re.search('CFL', s) or re.search('compact fluorescent.*', s) or re.search('energ.* conserve.*', s)):
        
        score += 1
        print('elc - energy manage hit')
    
    #building tech
    if ((re.search('build.*', s) or re.search('construct.*', s)) and (re.search('insula.*', s) or re.search('heat.* retent.*', s) 
        or re.search('heat.* exchang.*', s) or re.search('heat.* pump.*', s) or re.search('therm.* exchang.*', s) 
        or re.search('therm.* decompos.*', s) or re.search('therm.* energ.*', s) or re.search('therm.* communic.*', s) 
        or re.search('thermoplast.*', s) or re.search('thermocoup.*', s) or re.search('heat.* recover.*', s))):
        
        score += 1
        print('elc - building tech hit')
    
    return score

for cell in ws['D']:
    row = cell.row
    print(row)
    
    if row != 1:
        general_score = general(str(cell.value))
        environ_score = environ(str(cell.value))
        renew_score = renew(str(cell.value))
        elc_score = elc(str(cell.value))
        
        ws.cell(row=cell.row, column=5).value = general_score
        ws.cell(row=cell.row, column=6).value = environ_score
        ws.cell(row=cell.row, column=7).value = renew_score
        ws.cell(row=cell.row, column=8).value = elc_score
    
wb.save('editedawardscore.xls')